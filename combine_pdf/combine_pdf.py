import fnmatch
from openpyxl import load_workbook
from PyPDF2 import PdfFileReader, PdfFileWriter
import re
from sys import executable, path
import tkinter as tk
from tkinter import filedialog
from tkinter import messagebox
from tkinter.constants import N
import pandas as pd
import glob
from pandas.core.indexes.base import Index

from pandas.core.indexes.datetimes import date_range
import datetime
import time
import os
from escpos.connections import getFilePrinter, getNetworkPrinter
# from win32printing import Printer
import win32print
import win32api
import tempfile
import socket
import win32con
# import win32ui
# from win32printing import Printer
from win32printing import PrinterBase, Printer
# from PDFNetPython3 import *
import openpyxl


import PyPDF2
import re

# open the pdf file
# object = PyPDF2.PdfFileReader("111878_M03300004301_10122020_Signed.pdf")

# # get number of pages
# NumPages = object.getNumPages()

# files = "11_Dec-2020 Speed Post_Data.xlsx"
# filename = pd.read_excel(files)
# cust_id = filename["Customer ID"].tolist()
# mon = filename["Month"].tolist()
# list_string = map(str, cust_id)
# final_cust_id = list(list_string)
# for c in final_cust_id:
#     print (c)

# # define keyterms
# # print(final_cust_id)
# # String = "111878"
# # extract text and do the search
#     for i in range(0, NumPages):
#         PageObj = object.getPage(i)

#         # print("this is page " + str(i))
#         Text = PageObj.extractText()
#         # print(Text)
#         ResSearch = re.search(c, Text)
#     # print(ResSearch)


# pdfFileObj = open('111878_M03300004301_10122020_Signed.pdf', 'rb')
# pdfReader = PyPDF2.PdfFileReader(pdfFileObj, strict=False)

# search_word = "111878"
# search_word_count = 0

# for pageNum in range(0, pdfReader.numPages):
#     pageObj = pdfReader.getPage(pageNum)
#     text = pageObj.extractText().encode('utf-8')
#     search_text = text.lower().split()
#     for word in search_text:
#         if search_word in word.decode("utf-8"):
#             search_word_count += 1

# print("The word {} was found {} times".format(search_word, search_word_count)
# root = tk.Tk()
# root.title("Excel Merge")
# canvas1 = tk.Canvas(root, width=700, height=700,
#                     bg='lightsteelblue2', relief='raised')
# canvas1.pack()

# _printer = StringVar(root)

# # canvas1.attributes('-fullscreen', True)


# def sel_printer(*args):
#     print(_printer.get())


# def PrintAction(event=None):

#     PRINTER_DEFAULTS = {"DesiredAccess": win32print.PRINTER_ALL_ACCESS}
#     pHandle = win32print.OpenPrinter(_printer.get(), PRINTER_DEFAULTS)
#     properties = win32print.GetPrinter(pHandle, 2)
#     properties['pDevMode'].Color = 1 if str(_color.get()) == "Color" else 2
#     properties['pDevMode'].Copies = 1
#     win32print.SetPrinter(pHandle, 2, properties, 0)

#     if not _filename:
#         messagebox.showerror("Error", "No File Selected")
#         return
#     elif not _printer.get():
#         messagebox.showerror("Error", "No Printer Selected")
#         return

#     try:
#         # win32print.SetDefaultPrinter(_printer.get())
#         win32api.ShellExecute(0, "print", _filename, None,  ".",  0)
#         win32print.ClosePrinter(pHandle)
#     except:
#         pass
#         messagebox.showerror(
#             "Error", "There was an error printing the file :(")


# choices = [printer[2] for printer in win32print.EnumPrinters(2)]
# _printer.set(win32print.GetDefaultPrinter())  # set the default option


# def get_excel():
#     global filename

#     import_directory_path = filedialog.askopenfilename()
#     # filename = glob.glob(import_directory_path + "\*.xlsx")
#     filename = pd.read_excel(import_directory_path)
#     # filename.append()
#     doc_id = filename["Document No"]
#     print(doc_id)


# browseButton_Excel = tk.Button(text="       Select Excel File       ",
#                                command=get_excel, bg='green', fg='white', font=('helvetica', 12, 'bold'))
# canvas1.create_window(350, 130, window=browseButton_Excel)


# def by_doc_id():
#     global filename, doc_id
#     doc_id = filename["Document No"]
#     print(doc_id)

#     # col = doc_id.iloc()
#     # print(col)
#     # list_store.append(doc_id)

#     # print(list_store)


# browseButton_Excel = tk.Button(text="       by_doc_id       ",
#                                command=by_doc_id, bg='green', fg='white', font=('helvetica', 12, 'bold'))
# canvas1.create_window(350, 180, window=browseButton_Excel)


# def find_pdf():
#     global filename, doc_id

#     import_directory_path = filedialog.askdirectory()
#     # print(import_directory_path)
#     # names = [os.path.basename(x) for x in glob.glob(
#     # import_directory_path + "\*[A-Z ][0-9]+.pdf")]
#     # print(names)

#     paths = glob.glob(import_directory_path + "/*[A-Z ][0-9]+.pdf")
#     print(paths)

#     # for i in names:
#     # print(i)
#     # doc = re.findall(r'[A-Z ][0-9]+', str(i))
#     # print(doc)

#     # for j in doc:
#     # if doc_id in j:
#     # print("yes")

#     # print(j)
#     # doc_id = j
#     # pass
#     # print(doc_id)


# browseButton_Excel = tk.Button(text="       find_pdf       ",
#                                command=find_pdf, bg='green', fg='white', font=('helvetica', 12, 'bold'))
# canvas1.create_window(350, 250, window=browseButton_Excel)

# root.mainloop()

# def merge_pdfs(paths, output):
#     pdf_writer = PdfFileWriter()

#     for path in paths:
#         pdf_reader = PdfFileReader(path)
#         for page in range(pdf_reader.getNumPages()):
#             # Add each page to the writer object
#             pdf_writer.addPage(pdf_reader.getPage(page))

#     # Write out the merged PDF
#     with open(output, 'wb') as out:
#         pdf_writer.write(out)

# if __name__ == '__main__':
#     paths = glob.glob("*.pdf")

#     merge_pdfs(paths, output='merged.pdf')

# # print(filename)
# # merger = PdfFileMerger()

# # print(merger)
# # for pdf in filename:
# #     pdfobject = open(pdf, 'rb')
# #     merger.append(pdfobject)
# #     pdfobject.close()

# # merger.write("result.pdf")
# # merger.close()
# import socket
# mysocket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
# host = "  .168.2.149"
# port = 9100
# try:
#     soc = mysocket.connect((host, port))  # connecting to host
#     print(soc)
# # mysocket.send(b"^XA^A0N,50,50^FO50,50^FDSocket Test^FS^XZ")#using bytes
#     # mysocket.close()  # closing connection
# except:
#     print("Error with the E:\python\data\combine_pdf

# A List containing the system printers
# all_printers = [printer[2] for printer in win32print.EnumPrinters(2)]
# # Ask the user to select a printer
# printer_num = int(input("Choose a printer:\n" +
#                         "\n".join([f"{n} {p}" for n, p in enumerate(all_printers)])+"\n"))
# # set the default printer
# win32print.SetDefaultPrinter(all_printers[printer_num])
# pdf_dir = "E:\python\data\combine_pdf\*.pdf"
# # filename = "122940_SEP'2020_Signed.pdf"

# # print(pdf_dir)
# for f in glob(pdf_dir):
#     time.sleep(5)

#     print(f)
#     win32api.ShellExecute(0, "print", f, None,  ".",  0)

# input("press any key to exit")
# filename = glob.glob("*.pdf")
# print(filename)
# filename = "122940_SEP'2020_Signed.pdf"
# print(filename)
# for file in filename:

# time.sleep(5)
# os.startfile(file, "print")

# printer = getNetworkPrinter()(host='192.168.2.149', port=9100)

# printer.lf()

# def find_files(filename, search_path):
#     result = []

# Wlaking top-down from the root
# for root, dir, files in os.walk(search_path):
#     if filename in files:
#         result.append(os.path.join(root, filename))
# return result

# print(find_files(glob.glob("*.pdf"), "E:\python\data\combine_pdf"))
# all_data = pd.DataFrame()

# print("yes")

# print(pdf_dir)
# for f in glob(pdf_dir):

#     print(f)

# writer = pd.ExcelWriter(path, engine='xlsxwriter')

# ws = pd.to_excel(writer, sheet_name=None, index=False)
# print(ws)
# writer.save()

# print(doc_name)
# print(i)

# print(filename)
# filename.to_excel("new_excel.xlsx", index=False)
# print(jsonF)
# time.sleep(2)
# path2 = "new_excel.xlsx"
# filename = pd.read_excel(path2)

# all_data.append(filename)
# df = filename.style.index()
# print(all_data.reset_index(drop=True))
# print(all_data)
# print(filename)

# The source xlsx file is named as source.xlsx
# wb = load_workbook(files)

# ws = wb.active

# first_column = ws['C']
# import glob\

# print("Yess")

# print("yess " + i)
# return True

# listOfFiles = os.listdir('.')
# pattern = "*.pdf"
# for entry in listOfFiles:
#     if fnmatch.fnmatch(entry, pattern):
#         print(entry)
# filename = "122940_SEP'2020_Signed.pdf"
# for dirpath, dirs, files in os.walk(path):
# for filename in dirs:
# pass
# print(filename)
# fname = os.path.join(dirs, filename)
# if fname.endswith('.pdf'):
# print(fname)
# path = "E:\python\data\combine_pdf"
# filenames = glob.glob('[A-Z ][0-9]+*.pdf')
# print(filenames)
# for name in glob.glob('*[A-Z ][0-9]+*.pdf'):
# print(name)
# # Print the contents
# for x in range(len(first_column)):
# doc_no = first_column[x].value
# print(doc_no)

#   for root, dirs, files in os.walk(path):
# print(doc_no)
# if doc_no in files:
# print("True")
# else:
# print("hello")

# def find_all(name, path):
# result = []

# print(files)
# if name in files:
# result.append(os.path.join(root, name))
# print(result)
# return result

# find_all("106398_M05300009414_10122020_Signed.pdf",
#  "E:\python\data\combine_pdf")

# if doc_no in

# dis = filename.to_dict()
# doc_id = filename["Document No"]
# print(dis["Document No"])
# filename.dropna()

# Doc = dis["Document No"]

# print(doc_id)
# font = {
# "height": 8,
# }
# filename = "122940_SEP'2020_Signed.pdf"

# printer.text()
# Printer(doc_name=filename)
# print(PrinterBase.get_default_doc_name(filename))
# with Printer() as printer:
# printer.start()


# pass
# printer.start_doc()
# printer.text("title1", font_config=font)
# printer.text("title2", font_config=font)
# printer.text("title3", font_config=font)
# printer.text("title4", font_config=font)
# printer.new_page()
# printer.text("title5", font_config=font)
# printer.text("title6", font_config=font)
# printer.text("title7", font_config=font)
# printer.text("title8", font_config=font)
# print(Printer.get_default_doc_name())

# print(printer)
# filename = glob.glob("*.pdf")
# printer = win32print.GetDefaultPrinter()
# PRINTER_DEFAULTS = {"DesiredAccess": win32print.PRINTER_ALL_ACCESS}
# pHandle = win32print.OpenPrinter(printer, PRINTER_DEFAULTS)
# properties = win32print.GetPrinter(pHandle, 2)
# properties['pDevMode'].Color = 2
# properties['pDevMode'].PaperSize = 0
# properties['pDevMode'].PaperWidth = 210
# properties['pDevMode'].PaperLength = 297

# print(color)
# win32print.SetPrinter(pHandle, 2, properties, 0)
# print(p)
# win32api.ShellExecute(0, "print", filename, None,  ".",  0)
# win32print.ClosePrinter(pHandle)
# host = "192.168.2.149"

# printer = win32print.AddPrinterConnection()

# print(printer)
# printer = win32print.GetDefaultPrinter()
# PRINTER_DEFAULTS = {"DesiredAccess": win32print.PRINTER_ALL_ACCESS}
# pHandle = win32print.OpenPrinter(printer, PRINTER_DEFAULTS)
# level = 2
# properties = win32print.GetPrinter(pHandle, level)
# print(properties)

# pDevModeObj = properties["pDevMode"]
# pDevModeObj.PaperSize = 0
# pDevModeObj.PaperLength = 5334  # SIZE IN 1/10 mm
# pDevModeObj.PaperWidth = 7543  # SIZE IN 1/10 mm
# properties["pDevMode"] = pDevModeObj
# # win32print.SetPrinter(pHandle, level, properties, 0)
# # win32api.ShellExecute(0, "printto", filename, '"%s"' % printer, ".", 0)
# # win32print.ClosePrinter(pHandle)

# mysocket = socket.socket(socket.AF_INET, socket.SOCK_STREAM)
# host = "192.168.2.149"
# port = 9100
# # try:
# mysocket.connect((host, port))  # connecting to host
# filename = "new_files/122940_SEP'2020_Signed.pdf"
# # print(filename)
# mysocket.send(byte(filename))  # using bytes

# # print(PRINTER_DEFAULTS)
# pHandle = win32print.OpenPrinter(printer, PRINTER_DEFAULTS)

# print(pHandle)

# mysocket.send(filename)  # using bytes
# print("connected")
# mysocket.close()  # closing connection
# except:
# print("Error with the connection")
